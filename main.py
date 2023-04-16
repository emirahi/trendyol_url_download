import openpyxl
import sqlite3
import re
import bs4
import requests
import json

DATABASE_NAME = "database.db" # Veritabanının adı
EXCEL_FILE_PATH = input("dosya adını giriniz : ") # Excel dosyasının adı girdi olarak alınır.
EXCEL_OUTPUT_FILE_PATH = "output.xlsx" # Excel dosyasının çıktı olarak alınacağı adı girdi olarak alınır.
DATABASE_PATH = "database.db" # Veritabanının adı  girdi olarak alınır.
SHEET_NAME = "Sheet1" # Excel dosyasındaki sayfanın adı


def unique_list(list1):
    """
    Verilen listenin içindeki tekrar eden elemanları çıkarır.
    """
    unique_list = []
    for x in list1:
        if x not in unique_list:
            unique_list.append(x)
    return unique_list


def pull_url_number_from_db(database_path):
    """
    -p- ve -? arasındaki ıniq sayıları çeker.
    """
    number_list=[]
    for i in read_sqlite(database_path):
        if i is None:
            continue
        else :
            try:
                number_list.append(re.findall(r"(?<=\-)[0-9]+(?=\?)",i)[0])
            except IndexError:
                number_list.append(re.findall(r"(?<=\-)[0-9]+",i)[0])
    return unique_list(number_list)


def pull_url_number_from_excel(excel_file_path):
    """
    -p- ve -? arasındaki uniq sayıları çeker.
    """
    number_list=[]
    for i in read_excel(excel_file_path):
        if i is None:
            continue
        else:
            print(i)
            try:
                number_list.append(re.findall(r"(?<=\-)[0-9]+(?=\?)",i)[0])
            except IndexError:
                number_list.append(re.findall(r"(?<=\-)[0-9]+",i)[0])
    return unique_list(number_list)


def url_diff_db(excel_data, number_list):
        """
        Excel dosyasındaki verileri okur database ile karşılaştırır ve olmayanları ekler.
        """
        excel_output=[]
        for i in excel_data:
            if i is not None :
                try:
                    if re.findall(r"(?<=\-)[0-9]+(?=\?)",i)[0] not in number_list:
                        details=get_details(i)
                        insert_product_details(details)
                        excel_output.append(details)
                except IndexError:
                    if re.findall(r"(?<=\-)[0-9]+",i)[0] not in number_list:
                        details=get_details(i)
                        insert_product_details(details)
                        excel_output.append(details)
        return excel_output

def insert_product_details(product_details):
    """
    Ürün detaylarını veritabanına kaydeder.
    """
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO product_details VALUES (?,?,?,?,?,?)", (product_details["title"],product_details["description"],product_details["price"],product_details["image"],product_details["url"],product_details["barcode"]))
    conn.commit()
    conn.close()


def create_database():
    """
    Veritabanı oluşturur.
    """
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS product_details (title TEXT, description TEXT, price TEXT, image TEXT, url TEXT, barcode TEXT)")
    conn.commit()
    conn.close()


def read_excel(excel_file_path):
    """
    Excel'den url'leri çeker.
    """
    excel_data = []
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[SHEET_NAME]
    for i in range(1, sheet.max_row + 1):
        excel_data.append(sheet.cell(row=i, column=1).value)
    return excel_data


def read_sqlite(sqlite_path):
    """
    Veritabanından url'leri çeker.
    """
    conn = sqlite3.connect(sqlite_path)
    cursor = conn.cursor()
    cursor.execute("SELECT url FROM product_details")
    sqlite_data = cursor.fetchall()
    conn.close()
    sqlite_url_data= [i[0] for i in sqlite_data]
    return sqlite_url_data


def get_details(url):
    """
    url'den ürün detaylarını çeker.
    """
    details_dict = {"title": "", "description": "", "price": "", "image": "", "url": "", "barcode": ""}
    response=requests.get(url=url)
    soup=bs4.BeautifulSoup(response.text,"html.parser")
    for i in soup.find_all("section",{"class":"details-section"}): 
        details_dict["title"] = i.find("h3", {"class":"detail-name"}).text
    for i in soup.find_all("section",{"class":"details-section"}):
        details_dict["description"] = i.find("ul", {"class":"detail-desc-list"}).text
    for i in soup.find_all("script",{"type":"application/javascript"}):
        if "window.__PRODUCT_DETAIL_APP_INITIAL_STATE__=" in i.text:
            data = json.loads(re.findall(r'window.__PRODUCT_DETAIL_APP_INITIAL_STATE__=({.*?});', i.text)[0])
            details_dict["price"] = data["product"]["variants"][0]["price"]["sellingPrice"]["value"] 
    for i in soup.find_all("div",{"class":"gallery-container"}):
        details_dict["image"] = i.find("img").get("src")
    for i in soup.find_all("script",{"type":"application/javascript"}):
        if "window.__PRODUCT_DETAIL_APP_INITIAL_STATE__=" in i.text:
            data = json.loads(re.findall(r'window.__PRODUCT_DETAIL_APP_INITIAL_STATE__=({.*?});', i.text)[0])
            details_dict["barcode"] = data["product"]["variants"][0]["barcode"]
    details_dict["url"] = url
    return details_dict


def url_diff_excel(excel_data, number_list):
    """
    Excel dosyasındaki verileri okur database ile karşılaştırır ve olmayanları ekler.
    """
    excel_urls=[]
    for i in excel_data:
        if i is not None and number_list is not None:
            try:
                if re.findall(r"(?<=\-)[0-9]+(?=\?)",i)[0] in number_list:
                    number_list.remove(re.findall(r"(?<=\-)[0-9]+(?=\?)",i)[0])
                    excel_urls.append(i)
            except IndexError:
                if re.findall(r"(?<=\-)[0-9]+",i)[0] in number_list:
                    number_list.remove(re.findall(r"(?<=\-)[0-9]+",i)[0])
                    excel_urls.append(i)
    return excel_urls


def create_excel_file(excel_file_path):
    """
    Excel dosyası oluşturur.
    """
    wb = openpyxl.Workbook()
    wb.create_sheet(SHEET_NAME)
    wb.save(excel_file_path)
    
def create_excel_sheet(excel_file_path, sheet_name):
    """
    Excel dosyasına sheet ekler.
    """
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        wb.create_sheet(sheet_name)
        wb.save(excel_file_path)
    except FileNotFoundError:
        create_excel_file(excel_file_path)
        wb = openpyxl.load_workbook(excel_file_path)
        wb.create_sheet(sheet_name)
        wb.save(excel_file_path)

def dictionary_to_excel(excel_file_path, dictionary):
    """
    Dictionary'yi excel dosyasına ekler.
    """
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[SHEET_NAME]
    for i in range(1, len(dictionary)+1):
        sheet.cell(row=i, column=1).value = dictionary[i-1]["title"]
        sheet.cell(row=i, column=2).value = dictionary[i-1]["description"]
        sheet.cell(row=i, column=3).value = dictionary[i-1]["price"]
        sheet.cell(row=i, column=4).value = dictionary[i-1]["image"]
        sheet.cell(row=i, column=5).value = dictionary[i-1]["url"]
        sheet.cell(row=i, column=6).value = dictionary[i-1]["barcode"]
    wb.save(excel_file_path)




create_database()

unique_excel_list=pull_url_number_from_excel(EXCEL_FILE_PATH) #excel dosyasındaki url'leri okur unique olarak
unique_sqlite_list=pull_url_number_from_db(DATABASE_PATH) #database dosyasındaki url'leri okur unique olarak 

excel_file=read_excel(EXCEL_FILE_PATH) #excel dosyasındaki url'leri okur

excel_data=url_diff_excel(excel_file, unique_excel_list) #excel dosyasındaki url'leri okur excel ile karşılaştırır -p- ? değerine göre
output=url_diff_db(excel_data, unique_sqlite_list) #excel dosyasındaki url'leri okur database ile karşılaştırır -p- ? değerine göre ardından database e ekler

create_excel_sheet(EXCEL_OUTPUT_FILE_PATH, SHEET_NAME) #excel dosyasına sheet ekler
dictionary_to_excel(EXCEL_OUTPUT_FILE_PATH,output) #excel dosyasına verileri yazar dışarı export eder





